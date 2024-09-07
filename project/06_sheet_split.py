import openpyxl
#from xlutils.copy import copy

def get_data():
    wb = openpyxl.load_workbook('./base_data/data01.xlsx')
    sh = wb.worksheets[0]

    all_data = {}
    for r in sh.iter_rows(min_row=1, values_only=True):
        d = {'type': r[1], 'name': r[2], 'count': r[3], 'price': r[4]}
        key = r[0]
        if all_data.get(key):
            all_data[key].append(d)
        else:
            all_data[key] = [d]
    return all_data




def save(data):
    wb = openpyxl.load_workbook('./base_data/data01.xlsx')
    for key in data.keys():
        temp_sheet = wb.create_sheet(key)
        for i, d in enumerate(data.get(key), start=1):
            temp_sheet.cell(i,1,d.get('type'))
            temp_sheet.cell(i,2,d.get('name'))
            temp_sheet.cell(i,3,d.get('count'))
            temp_sheet.cell(i,4,d.get('price')) 
    wb.save('./generate_data/06_sheet_split.xlsx')



if __name__ == "__main__":
    all_data = get_data()
    #print(all_data)
    save(all_data)