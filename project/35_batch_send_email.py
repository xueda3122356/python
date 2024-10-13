import smtplib
from openpyxl import load_workbook
from email.mime.text import MIMEText
from email.header import Header


# 登录邮箱
smtp_obj = smtplib.SMTP("smtp.gmail.com", 587)
smtp_obj.starttls()  # 启动 TLS 加密
smtp_obj.login('xueda3122356@gmail.com', 'jishtrpmnetbfhvl')

# 读取excel内容获取工资数据
wb = load_workbook('./base_data/工资数据.xlsx')
sh =wb.active

# 创建工资表
table_header = '<tr>'
for index, row in enumerate(sh.iter_rows()):
    if index == 0:
        for col in row:
            table_header += f'<td>{col.value}</td>'
        table_header += '</tr>'
        continue
    else:
        table_info = '<tr>'
        for col in row:
            table_info += f'<td>{col.value}</td>'
        table_info += '</tr>'
    name = row[1].value
    address = row[9].value


    text_body = f'''
    <h3>您好，{name}</h3>
    <p>请查收2024年9月的工资详情:</p>
    <table border="1">
    {table_header}
    {table_info}
    </table>
'''

# 封装邮件
    msg_body = MIMEText(text_body, "html", 'utf-8')
    msg_body['From'] = Header("人事部", 'utf-8')
    msg_body['Subject'] = Header("启航2024年9月工资条", 'utf-8')

    smtp_obj.sendmail('xueda3122356@gmail.com', ['xueda3122356@hotmail.com'], msg_body.as_string())
    print(f'成功发送邮件给{name},邮箱地址是{address}')
