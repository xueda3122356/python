# pip install openpyxl

def open():
    from openpyxl import load_workbook
    wb = load_workbook('./base_data/data01.xlsx')
    sh1 = wb.active
    sh2 = wb.worksheets[0]
    sh3 = wb['Sheet1']
    sh4 = wb.get_sheet_by_name('Sheet1')
    print(sh1 is sh2 is sh3 is sh4)

def show_sheets():
    from openpyxl import load_workbook
    wb = load_workbook('./base_data/data01.xlsx')
    print(wb.sheetnames)

    for i in wb:
        print(i.title)

def get_one_value():
    from openpyxl import load_workbook
    wb = load_workbook('./base_data/data01.xlsx')
    sh = wb.active

    cell_value = sh.cell(2,3).value
    cell_value2 = sh['c2'].value

    print(cell_value, cell_value2)

def get_many_values():
    from openpyxl import load_workbook
    wb = load_workbook('./base_data/data01.xlsx')
    sh = wb.active

    # 切片
    cells_value = sh['c2': 'd3']

    print(cells_value)

    # 整行，整列
    cells_row3 = sh[3]
    cells_col3 = sh['c']
    cells_row3_5 = sh[3:5]
    cells_col3_5 = sh['c':'e']
    print(cells_row3)
    print(cells_col3)
    print(cells_row3_5)
    print(cells_col3_5)

    # 通过iterative获取数据
    # 通过获取整行数据进行iterative
    for row in sh.iter_rows(min_row=2, max_row=4, min_col=2):
        for cell in row:
            print(cell.value)
    # 通过获取整列数据进行iterative
    for row in sh.iter_cols(min_row=2, min_col=2, max_col= 4):
        for cell in row:
            print(cell.value)

def get_all_data():
    from openpyxl import load_workbook
    wb = load_workbook('./base_data/data01.xlsx')
    sh = wb.active

    for row in sh.rows:
        for cell in row:
            print(cell.value)

def get_num():
    from openpyxl import load_workbook
    wb = load_workbook('./base_data/data01.xlsx')
    sh = wb.active

    num_rows = sh.max_row
    num_cols = sh.max_column

    print(num_rows, num_cols)

if __name__ == "__main__":
    #open()
    #show_sheets()
    #get_one_value()
    #get_many_values()
    #get_all_data()
    get_num()