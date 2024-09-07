def new():
    from openpyxl import Workbook
    wb = Workbook()
    sh1 = wb.active
    sh3 = wb.create_sheet('image')
    sh2 = wb.create_sheet('data', 0)

    wb.save('./generate_data/08_create_workbook_and_sheet.xlsx')

def set_value():
    from openpyxl import Workbook
    from openpyxl.styles import Font, colors
    bold_italic_30_font = Font(name='Times New Roman', size=30, italic=True, bold=True, color=colors.BLUE)
    wb = Workbook()
    sh1 = wb.active
    sh1['A1'] = 'Hello world!'
    sh1['A1'].font = bold_italic_30_font
    wb.save('./generate_data/08_set_value.xlsx')

def set_value2():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    data = ['python','java','c++']

    for i, d in enumerate(data):
        #ws.cell(i+1, i+1, d)
        ws.cell(i+1, i+1).value = d
    
    wb.save('./generate_data/08_set_value2.xlsx')

def set_style():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    wb = Workbook()
    ws = wb.active
    data = ['python','java','c++']
    ws.row_dimensions[1].height = 30
    ws.column_dimensions['A'].width = 20

    for i, d in enumerate(data):
        #ws.cell(i+1, i+1, d)
        ws.cell(i+1, i+1).value = d
        ws.cell(i+1, i+1).alignment = Alignment(horizontal='right', vertical='top')
        
    wb.save('./generate_data/08_set_style.xlsx')


def set_merge():
    from openpyxl import Workbook
    wb = Workbook()
    sh1 = wb.active
    sh1.merge_cells('a1:c1') # 单元格合并时，只能从小到大，从左到右，从上到下
    sh1.merge_cells('d2:e5')
    sh1['a1'] = '横向合并' # 单元格合并时，只有第一个单元格可以读写
    sh1['d2'] = '多合并'

    wb.save('./generate_data/08_cells_merge.xlsx')

def set_img():
    from openpyxl import Workbook
    wb = Workbook()
    sh1 = wb.active
    from openpyxl.chart import LineChart
    from datetime import date
    rows = [
        ['Date', 'Batch 1', 'Batch 2', 'Batch 3'],
        [date(2024,9, 1), 40, 30, 50],
        [date(2024,9, 2), 40, 25, 50],
        [date(2024,9, 3), 40, 30, 50],
        [date(2024,9, 4), 35, 30, 50],
        [date(2024,9, 5), 40, 25, 50],
        [date(2024,9, 6), 40, 30, 45],
        [date(2024,9, 7), 30, 35, 50],
    ]

    for d in rows:
        sh1.append(d)
    from openpyxl.chart import Reference
    c1 = LineChart()
    c1.title = 'Line Chart'
    c1.x_axis.title = 'Test_number'
    c1.y_axis.title = 'Size'
    # min_col:从第几列开始取值， min_row: 从第几行开始取值， max_col: 取值到第几列， max_row: 取值到第几行
    data = Reference(sh1, min_col = 2, min_row = 1, max_col = 4,max_row = 8)
    c1.add_data(data, titles_from_data = True)
    sh1.add_chart(c1, 'a12')
    wb.save('./generate_data/08_set_img.xlsx')

def set_img2():
    from openpyxl import Workbook
    # Create a new workbook and add data
    wb = Workbook()
    sh1 = wb.active
    rows = [
        ['名称', '数值'],
        ['苹果', '0.5'],
        ['香蕉', '0.2'],
        ['西瓜', '0.15'],
        ['樱桃', '0.15'],
    ]
    
    for d in rows:
        sh1.append(d)
    
    # Create a pie chart
    from openpyxl.chart import PieChart, Reference
    c1 = PieChart()
    c1.title = "Fruit Data"  # Set the chart title

    # Reference for labels and data (exclude the header row)
    labels = Reference(sh1, min_col=1, min_row=2, max_row=5)  # Labels from row 2 to 5
    data = Reference(sh1, min_col=2, min_row=2, max_row=5)  # Data from row 2 to 5
    
    c1.add_data(data, titles_from_data=False)  # Add data without titles
    c1.set_categories(labels)  # Set the categories for labels
    
    # Add the chart to the sheet at a new position
    sh1.add_chart(c1, 'E5')
    
    # Save the workbook
    wb.save('./generate_data/08_set_img2.xlsx')


def set_img3():
    from openpyxl import Workbook
    # Create a new workbook and add data
    wb = Workbook()
    sh1 = wb.active
    rows = [
        ['Number', 'Batch 1', 'Batch 2'],
        [2, 10, 30],
        [3, 50, 60],
        [4, 20, 70],
        [5, 10, 10],
        [6, 50, 40],
        [7, 40, 30],
        [8, 10, 20],
    ]

    for row in rows:
        sh1.append(row)

    from openpyxl.chart import BarChart, Reference
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 4
    chart1.title = "Bar Chart"
    chart1.y_axis_title = "Test Number"
    chart1.x_axis_title = "Sample Length (mm)"

    data = Reference(sh1, min_col=2, min_row=1, max_col=3, max_row=8)
    cats = Reference(sh1, min_col=1, min_row=2, max_row=8)

    chart1.add_data(data, titles_from_data = True)
    chart1.set_categories(cats)
    chart1.style = 10
    sh1.add_chart(chart1, "f5")
    wb.save('./generate_data/08_set_img3.xlsx')

if __name__ == "__main__":
    #new()  
    #set_value()
    #set_value2()
    #set_style()
    #set_merge()
    #set_img()
    #set_img2()
    set_img3()