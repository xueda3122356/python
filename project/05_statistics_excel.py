# pip install openpyxl
import openpyxl
from xlutils.copy import copy

def read_data():
    wb = openpyxl.load_workbook('./base_data/data01.xlsx', data_only= True)
    sh = wb.worksheets[0]

    catogory_type = {}  # {a:100, b:200}
    count_total = []    # [1,2,3,4,5,6]

    for r in sh.iter_rows(min_row=1,values_only=True):
        count = r[3] * r[4]
        count_total.append(count)

        key = r[0]
        if catogory_type.get(key):
            catogory_type[key] += count
        else:
            catogory_type[key] = count
    
    return catogory_type,count_total # total value of each company and total value of each item

def save_data(catogory, count):
    wb = openpyxl.load_workbook('./base_data/data01.xlsx') 
    sh_rd = wb.worksheets[0]

    newColumn = sh_rd.max_column + 1
    for r, row in enumerate(sh_rd.iter_rows(min_row=1), start=0):
        sh_rd.cell(row=r+1,column=newColumn,value=count[r])
    
    sh2 = wb.create_sheet('汇总金额')

    for i,key in enumerate(catogory.keys(), start=1):   
        sh2.cell(i,1,key)
        sh2.cell(i,2,catogory.get(key))

    wb.save('./generate_data/05_statistics_excel.xlsx')


if __name__ == "__main__":
    catogory,count = read_data()
    save_data(catogory= catogory, count= count)


