from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def dum():
    wb = load_workbook('./base_data/worktime.xlsx')
    sh = wb.active

    index = []
    temp = []

    for i, c in enumerate(sh['B']):
        if c.value not in temp:
            temp.append(c.value)
            print(i)
        else:
            index.append(i)
        
    for i, r in enumerate(sh.rows):
        if i in index:
            for c in r:
                c.fill = PatternFill('solid', fgColor= 'AEEEEE')
            print(i)
    wb.save('./generate_data/mark_duplicate_data.xlsx')

if __name__ == "__main__":
    dum()