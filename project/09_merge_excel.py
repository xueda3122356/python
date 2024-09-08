from openpyxl import load_workbook, Workbook
import os

def copy_data():
    wb = Workbook()
    sh = wb.active

    all_data = []

    for name in os.listdir('./sell_sheets'):
        path = f'./sell_sheets/{name}'
        temp_wb = load_workbook(path)
        temp_sh = temp_wb.active
        for row in range(1, temp_sh.max_row+1):
            # aquire the value of the whole row
            row_value = []
            for col in range(1, temp_sh.max_column+1):
                value = temp_sh.cell(row,col).value
                row_value.append(value)
            # save the whole row's value to list
            if row_value not in all_data: # avoid the same data
                all_data.append(row_value)
    
    for d in all_data:
        sh.append(d)
    
    wb.save('./generate_data/09_merge_excel_data.xlsx')

if __name__ == "__main__":
    copy_data()