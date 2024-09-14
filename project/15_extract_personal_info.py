from openpyxl import load_workbook
from datetime import datetime

def extract_info():
    wb = load_workbook('./base_data/personal_information.xlsx')
    sh = wb.active

    now_year = datetime.now().year # get the current year
    max_column = sh.max_column

    for i,cell in enumerate(sh['b']):
        pno = cell.value
        year = pno[6:10] # 6位行政划区 4位年份 2位月份 2位日期 4位个人编码
        month = pno[10:12]
        day = pno[12:14]
        # calculate person's age
        age = now_year - int(year)
        sh.cell(i+1, max_column+1).value = year
        sh.cell(i+1, max_column+2).value = month
        sh.cell(i+1, max_column+3).value = day
        sh.cell(i+1, max_column+4).value = age
    wb.save('./generate_data/15_extract_personal_info.xlsx')

if __name__ == "__main__":
    extract_info()