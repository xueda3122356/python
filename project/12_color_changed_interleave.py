from openpyxl import Workbook
from openpyxl.styles import PatternFill


def create_excel():
    wb = Workbook()
    sh = wb.active

    rows = [
        ['Number', 'Batch 1', 'Batch 2'],
        [2, 10, 30],
        [3, 50, 60],
        [4, 20, 70],
        [5, 10, 10],
        [6, 50, 40],
        [7, 40, 30],
        [8, 10, 20],
    ]

    for row in rows:
        sh.append(row)

    bg_color = PatternFill('solid', fgColor= 'AEEEEE')

    # 隔行变色，偶数行变色
    for row in range(1, sh.max_row+1):
        if row % 2 == 0:
            for col in range(1, sh.max_column+1):
                sh.cell(row, col).fill = bg_color # 填充背景颜色

    wb.save('./generate_data/color_fill_interleave.xlsx')

if __name__ == "__main__":
    create_excel()
