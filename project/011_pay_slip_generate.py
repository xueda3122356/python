from openpyxl import load_workbook, Workbook

def create_paySlip():
    wb = load_workbook('./base_data/工资数据.xlsx')
    sh = wb.active

    title = []

    # 提取首行
    for col in range(sh.max_column):
        cell_value = sh.cell(1,col+1).value
        title.append(cell_value)
    
    # 创建excel单独存储个人工资条
    for i, row in enumerate(sh.rows):
        if i == 0:
            continue
        else:
            temp_wb = Workbook()
            temp_sh = temp_wb.active
            temp_sh.append(title)
            row_value = [cell.value for cell in row] # 遍历一行的信息
            temp_sh.append(row_value)
            temp_wb.save(f'./generate_data/{row_value[1]}.xlsx')

    
if __name__ == "__main__":
    create_paySlip()