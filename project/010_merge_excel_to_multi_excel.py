from openpyxl import load_workbook, Workbook
import os

def copy_data():
    wb = Workbook()

    for name in os.listdir('./sell_sheets'):
        path = f'./sell_sheets/{name}'
        temp_wb = load_workbook(path)
        temp_sh = temp_wb.active
        sh = wb.create_sheet(name[:-5])
        for row in range(1, temp_sh.max_row+1):
            # aquire the value of the whole row
            row_value = []
            for col in range(1, temp_sh.max_column+1):
                value = temp_sh.cell(row,col).value
                row_value.append(value)
            # save the whole row's value to sheet
            sh.append(row_value)
    del wb['Sheet']
    
    wb.save('./generate_data/010_merge_excels_data_to_multi_sheets.xlsx')

if __name__ == "__main__":
    copy_data()