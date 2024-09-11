from openpyxl import Workbook, load_workbook
from datetime import date

def create_data():
    wb = Workbook()
    sh = wb.active

    data = [
    ['Date', '姓名', '打卡时间'],
    ["2020-12-01", "吕小布", "18:50"],
    ["2020-12-02", "强的弹", "18:10"],
    ["2020-12-03", "刘鑫", "18:02"],
    ["2020-12-04", "吕小布", "18:50"],
    ["2020-12-05", "张飞", "19:22"],
    ["2020-12-06", "吕小布", "18:50"]
    ]

    for d in data:
        sh.append(d)
    
    wb.save('./base_data/worktime.xlsx')

def statistics_over_time():
    wb = load_workbook('./base_data/worktime.xlsx')
    sh = wb.active

    # 读取数据
    data = []
    for i in range(2, sh.max_row+1):
        t_value = []
        for j in range(1, sh.max_column+1):
            t_value.append(sh.cell(i,j).value)
        # 统计时间
        h,m = t_value[2].split(":")
        full_time = int(h)*60 + int(m)
        over_time = full_time - 18*60
        t_value.append(over_time)
        data.append(t_value)

    # 保存excel
    temp_wb = Workbook()
    temp_sh = temp_wb.active

    for d in data:
        temp_sh.append(d)

    temp_wb.save('./generate_data/13_statistics_over_time.py.xlsx')

if __name__ == "__main__":
    #create_data()
    statistics_over_time()